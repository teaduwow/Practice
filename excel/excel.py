from types import DynamicClassAttribute
from openpyxl import Workbook, load_workbook, workbook
from openpyxl.reader import excel
from openpyxl.worksheet import worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# wb = load_workbook('excel\excel.xlsx')
# ws = wb.active
# print(ws)
# print (ws['A5'].value)

# ws['A5'].value = "小灰"
# wb.save('excel\excel.xlsx')

# ws = wb["工作表2"]

# wb.create_sheet("qq")
# print (ws)
# # print(wb.sheetnames)

# # wb.save('excel\excel.xlsx')

# # 創建excel檔案
# wb = load_workbook("excel/3excel.xlsx")
# ws = wb.active
# # ws.title = "qq" 
# # ws.append([123,456,789]) #創建橫排資料
# # ws.append([123,456,785])
# # ws.append([561,154,156])


# # wb.save ('excel/3excel.xlsx')

# # for row in range(1,4):
# #     for col in range(1,4):
# #         char = get_column_letter(col)
# # ws[char + str(row)].value = char + str(row)

# # wb.save("excel/3excel.xlsx")
# # ws.merge_cells("A1:E2") #合併儲存格
# # ws.unmerge_cells("A1:E2")#取消合併

# # wb.save("excel/3excel.xlsx")

# # 插入 column & row
# # ws.insert_rows(3)
# # # ws.insert_cols(2)
# # ws.delete_cols(2)
# # ws.delete_rows(2)


# # 移動資料
# ws.move_range("A3:E4",rows=2,cols=2)

# wb.save("excel/3excel.xlsx")

data = [
    {
        "name": "白",
        "tall": 180,
        "age" : 23,
        "weight":74
    },
    {
        "name":"黑",
        "tall": 170,
        "age" : 45,
        "weight":50
    },
    {
        "name":"綠",
        "tall":180,
        "age":30,
        "weight":60,
    },
    {
        "name":"藍",
        "tall":195,
        "age":16,
        "wight":90
    },
    {
        "name":"黃",
        "tall":150,
        "age":16,
        "weught":45
    }
]
    


    
wb = Workbook()
ws = wb.active

title=["姓名","身高","年紀","體重"]
ws.append(title)

for person in data:
    ws.append (list(person.values()))

for col in range(2,5):
    char = get_column_letter(col)
    ws[char +"7"] = f'=average({char + "2"}:{char + "6"})'

for col in range(1,5):
    char = get_column_letter(col)
    ws[char +"1"].font = Font(bold=True, color="009999FF")


wb.save("excel/data.xlsx")



source = wb.active
target = wb.copy_worksheet(source)

for sheet in wb:
    print(sheet.title)
wb.save("excel/data.xlsx")
ws["A4"]= 4
d = ws.cell(row=4, column=2, value=10)

for x in range(1,101):
    for y in range(1,101):
        ws.cell(row=x, column=y)
cell_range = ws['A1':'C2'].value(12)


wb.save("excel/data.xlsx")

